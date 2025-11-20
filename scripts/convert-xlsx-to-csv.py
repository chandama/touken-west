#!/usr/bin/env python3
import sys
import csv
import openpyxl

def convert_xlsx_to_csv(xlsx_path, csv_path):
    """Convert Excel file to CSV"""
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheet = workbook.active

    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"Converted {xlsx_path} to {csv_path}")
    print(f"Sheet name: {sheet.title}")
    print(f"Rows: {sheet.max_row}, Columns: {sheet.max_column}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: convert-xlsx-to-csv.py <input.xlsx> <output.csv>")
        sys.exit(1)

    convert_xlsx_to_csv(sys.argv[1], sys.argv[2])
